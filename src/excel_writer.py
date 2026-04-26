import shutil
import os
import openpyxl as yxl
import re
import numpy as np
from src.core_logic import excelExtract, tableReformation, redispatch
from src.config import BOSTON_ZONES

def writeExcel(file, inputParameters):
    busNo, projectSizeMain, response = inputParameters
    
    outputFile = 'redispatch_' + os.path.basename(file)
    
    try:
        shutil.copyfile(file, outputFile)
    except Exception as e:
        raise Exception(f"Could not copy file (it may be open or permissions denied): {str(e)}")
        
    try:
        wb = yxl.load_workbook(outputFile)
        ws = wb.worksheets
    except Exception as e:
        raise Exception(f"Could not open workbook {outputFile}: {str(e)}")
   
    optimizationRange = list(range(0, projectSizeMain + 5, 5))
    optimizationRange.reverse()

    for sheet, sheetName in zip(ws, wb.sheetnames):
        try:
            flows, dispatch = excelExtract(file, sheetName)
        except Exception as e:
            # Skip sheets that don't match the format
            continue
            
        loading = 2.00
        for projectSize in optimizationRange:
            if loading > 1.02:
                if response in 'Yy':
                    charging = 'T'
                    projSize = 0 - projectSize
                elif response in 'Nn':
                    charging = 'F'
                    projSize = projectSize
                else:
                    charging = 'F'
                    projSize = projectSize
                    
                dispatch11 = dispatch.sort_values(by=['Dfax_1'], ascending=False)
                flow_data, dispatch1 = tableReformation(dispatch11, flows, busNo, projSize)
                
                zoneNumber_series = dispatch1.loc[dispatch1['  Bus#'] == busNo, 'Zone']
                zoneNumber = int(zoneNumber_series.iloc[0]) if not zoneNumber_series.empty else None

                switchN = 1 if zoneNumber in BOSTON_ZONES else 0

                redispatchFlows, dfDispatch = redispatch(flow_data, dispatch1, busNo, charging, switchN)
                
                redispatchFlows.replace(np.nan, '', inplace=True)
                dfDispatch.replace(np.nan, '', inplace=True)

                loading = max(redispatchFlows.iloc[:, -1])
                
                if loading > 1.02:
                    outputFile_intermediate = f"redispatch_{sheetName}_{projSize}MW_{file.split('/')[-1]}"
                    shutil.copyfile(file, outputFile_intermediate)
                    wb1 = yxl.load_workbook(outputFile_intermediate)
                    wsIndex = wb.sheetnames.index(sheetName)
                    ws1 = wb1.worksheets[wsIndex]

                    for r in range(3, 30):
                        if ws1.cell(row=r-2, column=1).value == 'Dispatch':
                            cnt = sum(1 for l in dfDispatch.columns if re.search('Dfax', str(l)))
                            colsOut = list(range(9)) + list(range(11, 11 + cnt))
                            for cols in colsOut:
                                for rows in range(r, len(dfDispatch) + r):
                                    ws1.cell(row=rows, column=cols+1).value = dfDispatch.iloc[rows-r, cols]
                    wb1.save(outputFile_intermediate)
                    wb1.close()
            else:
                for r in range(3, 30):
                    if sheet.cell(row=r-2, column=1).value == 'Dispatch':
                        cnt = sum(1 for l in dfDispatch.columns if re.search('Dfax', str(l)))
                        colsOut = list(range(9)) + list(range(11, 11 + cnt))
                        for cols in colsOut:
                            for rows in range(r, len(dfDispatch) + r):
                                sheet.cell(row=rows, column=cols+1).value = dfDispatch.iloc[rows-r, cols]
                wb.save(outputFile)
                wb.close()
                break

    return os.path.abspath(outputFile)
