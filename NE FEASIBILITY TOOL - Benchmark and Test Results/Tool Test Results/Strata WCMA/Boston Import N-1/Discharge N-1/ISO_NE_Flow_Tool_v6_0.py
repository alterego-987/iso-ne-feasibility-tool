# %%
import shutil
from urllib import response
import numpy as np
import openpyxl as yxl
import re
from csv import excel
from dis import dis
from fileinput import close
import os
from turtle import st
import pandas as pd

bostonZone = [1000, 1002, 1004, 1010, 1012, 1013, 1014, 1120, 1130]

def excelExtract(file,sheet):
    excelContent = pd.read_excel(file, sheet_name=sheet)#,skiprows=lambda x: x in [0,1,2,5,6,7])
    startingIndex_Flows = excelContent.index[excelContent.iloc[:,0]=='Flows'].tolist()
    startingIndex_Flows = startingIndex_Flows[0]+1
    startingIndex_Dispatch = excelContent.index[excelContent.iloc[:,0]=='Dispatch'].tolist()
    startingIndex_Dispatch = startingIndex_Dispatch[0]+1
    endingIndex_Flows = startingIndex_Dispatch-3

    flows = pd.DataFrame(
        excelContent.iloc[startingIndex_Flows+1:endingIndex_Flows,0:14],   
    )
    flows.columns = excelContent.iloc[startingIndex_Flows,0:14]

    dispatch = pd.DataFrame(excelContent.iloc[startingIndex_Dispatch+1:,:])
    dispatch.columns = columns=excelContent.iloc[startingIndex_Dispatch,:]
    dispatch.reset_index(drop=True)
    return flows, dispatch

# #function to calculate the dependent parameters after changing the values in the 'Pnew' column
def tableReformation(data,flowData, busNo, projectSize):
    #change in Pnew
    #print(projectSize)
    data.loc[data['  Bus#']==busNo,'Pnew']=projectSize

    #change in Pdelta
    pDelta = data.loc[data['  Bus#']==busNo,'Pnew']-data.loc[data['  Bus#']==busNo,'Pgen or Pload']
    data.loc[data['  Bus#']==busNo,'Pdelta']=pDelta
    
    #change in Pavail
    pAvail = data.loc[data['  Bus#']==busNo,'PMax Gen'] - data.loc[data['  Bus#']==busNo,'Pnew']
    data.loc[data['  Bus#']==busNo,'Pavail'] = pAvail
    
    #change in Impact_1
    #print(flowData.shape[0])
    for j in range(1,flowData.shape[0]+1):
        temp = 'Impact_'+str(j)
        #print(temp)
        impactVal = pDelta*data.loc[data['  Bus#']==busNo,'Dfax_'+str(j)]
        data.loc[data['  Bus#']==busNo, temp] = impactVal

    #change in FlowChange
    #print(flowData)
    for j in range(1,flowData.shape[0]+1):
        temp = 'Impact_'+str(j)
        #print(temp)
        flowChange = data[temp].sum()
        flowData.loc[j+2,'FlowChange'] = flowChange

    #change in FlowRes
        flowRes = flowData.loc[j+2,'FlowChange'] + flowData.loc[j+2,'FlowInit']
        flowData.loc[j+2,'FlowRes'] = flowRes

    #change in Loading
    #print(flowData.loc[:,'Limit'])
        flowData.loc[j+2,'Loading'] = flowData.loc[j+2,'FlowRes']/flowData.loc[j+2,'Limit']
    #print(flowData.loc[:,'Loading'])
    #a = max(flowData.loc[:,'Loading'])

    #print(flowData.loc[3,'FlowRes'])
    #print(data)
    return flowData, data

#function to perform re-dispatch
def redispatch(dispatchList,busNo,charging,switchNo=0):
    
    for i in range(dispatchList[1].shape[0]):
        dispatch = dispatchList[1]
        flows = dispatchList[0]
        pNew = dispatch.iloc[i].loc['Pnew']
        pmaxGen = dispatch.iloc[i].loc['PMax Gen']
        pavail = dispatch.iloc[i].loc['Pavail']
        busName = dispatch.iloc[i].loc['Bus Name    ']
        pDeltaSum = dispatch['Pdelta'].sum()

        # discharge
        #print(charging)
        if charging == 'F':
            #print('No')
            if pDeltaSum<=0:
                break
            else:
                if pNew == 0:
                    continue
                else:
                    if pNew > pDeltaSum:
                        droppingCapacity = pNew - pDeltaSum
                    elif pNew <=pDeltaSum:
                        droppingCapacity = 0
        
        elif charging == 'T':
            if pDeltaSum<=0:
                #if pNew == 0:
                if not pavail == 0:
                    if pmaxGen > pavail:
                        if abs(pDeltaSum)>pavail:
                            droppingCapacity = pNew + pavail
                        else:
                            droppingCapacity = pNew + abs(pDeltaSum)
                    elif pmaxGen <= pavail:
                        if abs(pDeltaSum)>pmaxGen:
                            droppingCapacity = pNew + pmaxGen
                        else:
                            droppingCapacity = pNew + abs(pDeltaSum)
                else:
                    continue
            else:
                break

        
        busN = dispatch.iloc[i].loc['  Bus#'] 
        zoneN = dispatch.loc[dispatch['  Bus#']==busN,'Zone']
        
        if not switchNo == 0: 
            if int(zoneN) in bostonZone:
                if not busN == busNo:
                    if 'SEABROOK' in busName:
                        pass
                    elif 'MILLSTONE' in busName:
                        pass
                    elif 'SUN' in busName:
                        pass
                    elif 'NYNE' in busName:
                        pass
                    elif 'NYPA' in busName:
                        pass
                    elif 'NBNE' in busName:
                        pass
                    else:
                        dispatchList = tableReformation(dispatch,flows,busN,droppingCapacity)
        else:
            if not busN == busNo:
                #print(busName)
                if 'SEABROOK' in busName:
                    pass
                elif 'MILLSTONE' in busName:
                    pass
                elif 'SUN' in busName:
                    pass
                elif 'NYNE' in busName:
                    pass
                elif 'NYPA' in busName:
                    pass
                elif 'NBNE' in busName:
                    pass
                else:
                    dispatchList = tableReformation(dispatch,flows,busN,droppingCapacity)
    
    pDeltaSum2 = dispatchList[1]['Pdelta'].sum()
    if not pDeltaSum2 <= 0:
        for i in range(dispatchList[1].shape[0]):
            pDeltaSum2 = dispatchList[1]['Pdelta'].sum()
            if not pDeltaSum2 <= 0:
                dispatch = dispatchList[1]
                flows = dispatchList[0]
                busName = dispatch.iloc[i].loc['Bus Name    ']
                pNew = dispatch.iloc[i].loc['Pnew']
                if pNew == 0:
                    continue
                else:
                    if pNew > pDeltaSum2:
                        droppingCapacity = pNew - pDeltaSum2
                    elif pNew <=pDeltaSum2:
                        droppingCapacity = 0
                    if not switchNo == 0:
                        busN = dispatch.iloc[i].loc['  Bus#'] 
                        zoneN = dispatch.loc[dispatch['  Bus#']==busN,'Zone']
                        dfaxN = dispatch.loc[dispatch['  Bus#']==busN,'Dfax_1']
                        if not int(zoneN) in bostonZone:
                            if -0.01 <= float(dfaxN) <= 0.01:
                                if not busN == busNo:
                                    if 'SEABROOK' in busName:
                                        pass
                                    elif 'MILLSTONE' in busName:
                                        pass
                                    elif 'SUN' in busName:
                                        pass
                                    else:
                                        dispatchList = tableReformation(dispatch,flows,busN,droppingCapacity)
            else:
                break
    return dispatchList

#funtion to write the final results in an excel file
def writeExcel(file, inputParameters):
    #print(wb.sheetnames)
    busNo = inputParameters[0]
    projectSizeMain = inputParameters[1]
    response = inputParameters[2]
    
    outputFile = 'redispatch_'+ file
    shutil.copyfile(file,outputFile)
    wb = yxl.load_workbook(outputFile)
    ws = wb.worksheets
   
    optimizationRange = list(range(0,projectSizeMain+5,5))
    list.reverse(optimizationRange)

    for sheet, sheetName in zip(ws,wb.sheetnames):
        dispatch = excelExtract(file, sheetName)
        loading = 2.00
        for projectSize in optimizationRange:
            if loading > 1.02:
                if response in 'Yy':
                    charging = 'T'
                    projSize = 0-projectSize

                if response in 'Nn':
                    charging = 'F'
                    projSize = projectSize
                    
                dispatch11 = dispatch[1].sort_values(by=['Dfax_1'],ascending=False)
                dispatch1 = tableReformation(dispatch11,dispatch[0],busNo,projSize)
                #print(dispatch1[0])
                zoneNumber = dispatch1[1].loc[dispatch1[1]['  Bus#']==busNo,'Zone']

                if int(zoneNumber) in bostonZone:
                    switchN = 1
                else:
                    switchN = 0

                redispatchResults = redispatch(dispatch1,busNo,charging,switchN)
                redispatchFlows = tableReformation(redispatchResults[1],redispatchResults[0],busNo,projSize)
                dfDispatch = redispatchResults[1]
                dfFlows = redispatchResults[0]
                dfFlows.replace(np.nan,'',inplace=True)
                dfDispatch.replace(np.nan,'',inplace=True)

                loading = max(redispatchFlows[0].iloc[:,-1])
                if loading > 1.02:
                    outputFile_intermediate = 'redispatch_'+ sheetName +'_' + str(projSize) + 'MW_' + file 
                    shutil.copyfile(file,outputFile_intermediate)
                    wb1 = yxl.load_workbook(outputFile_intermediate)
                    wsIndex = wb.sheetnames.index(sheetName)
                    #print(wb1.worksheets[0])
                    ws1 = wb1.worksheets[wsIndex]

                    for r in range(3,30):
                        if ws1.cell(row=r-2,column=1).value=='Dispatch':
                            cnt=0
                            for l in dfDispatch.columns:
                                if re.search('Dfax',l):
                                    cnt+=1
                            colsOut = list(range(9))+list(range(11,11+cnt))
                            for cols in colsOut:
                                for rows in range(r,len(dfDispatch)+r):
                                    ws1.cell(row=rows, column=cols+1).value = dfDispatch.iloc[rows-r,cols]
                    wb1.save(outputFile_intermediate)
                    wb1.close

                #print(loading)
                #print(redispatchResults[1]['Bus Name    '])
            else:
                for r in range(3,30):
                    if sheet.cell(row=r-2,column=1).value=='Dispatch':
                        cnt=0
                        for l in dfDispatch.columns:
                            if re.search('Dfax',l):
                                cnt+=1
                        colsOut = list(range(9))+list(range(11,11+cnt))
                        for cols in colsOut:
                            for rows in range(r,len(dfDispatch)+r):
                                sheet.cell(row=rows, column=cols+1).value = dfDispatch.iloc[rows-r,cols]
                wb.save(outputFile)
                wb.close
                break


if __name__=='__main__':
    # import cProfile, pstats
    # profiler = cProfile.Profile()
    # profiler.enable()

    file = 'Charging N-1 - Copy.xlsx' #input('Enter File Name with extension i.e., N-1_Trial.xlsx: \n')

    busNo = 999312 #int(input('Enter Bus Number for redispatch: \n'))
    projectSizeMain = 100 #float(input('Enter the project size: \n'))

    ips = (busNo, projectSizeMain, 'Y')
    writeExcel(file,ips)
    print('redispatch_' +file+' created!')
    # profiler.disable()
    # stats = pstats.Stats(profiler).sort_stats('cumtime')
    # stats.print_stats()
# %%
# %%
