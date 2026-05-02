"""
Generates a synthetic N-1 power flow study Excel file for demonstration.
All bus numbers, plant names, and transmission elements are fictional.
No CEII or client data is included.

The output file mirrors the exact visual structure of real TARA exports:
  - 5 monitored flows per sheet
  - 5 Dfax columns (Dfax_1 through Dfax_5)
  - Excel formulas for Pdelta, Pavail, Impact columns
  - SUM row before dispatch headers
  - Autofilter on dispatch header row
  - Color coding (red for binding constraint, yellow Pnew header, gray Pnew cells)

Run from the repo root:
    python sample_data/generate_sample.py

Suggested inputs when running the feasibility tool against the output:
    Bus Number  : 800001
    Project Size: 50
    Mode        : Discharging

Expected behavior:
    Sheet 1 - 50 MW raises Flow 1 loading to 1.032 (binding).
    Redispatch reduces ALPHA_GAS_1 + ALPHA_GAS_2 to bring loading to ~0.987.
    Sheet 2 - 50 MW raises max loading to 1.006. Already within limit.
"""

import os
import openpyxl as yxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── Column definitions ──────────────────────────────────────────────────────

FLOW_HEADERS = [
    "NN", "LFCaseName", "Monitored Element",
    None, None, None, None, None,
    "Contingency", "Limit", "FlowInit", "FlowRes", "FlowChange", "Loading",
]

DISPATCH_HEADERS = [
    "  Bus#", "Bus Name    ", "Volt", "Area", "Zone",
    "Pgen or Pload", "PMin Gen", "PMax Gen", "Pnew",
    "Pdelta", "Pavail",
    "Dfax_1", "Dfax_2", "Dfax_3", "Dfax_4", "Dfax_5", "---",
    "Impact_1", "Impact_2", "Impact_3", "Impact_4", "Impact_5",
]

# ─── Synthetic generators ────────────────────────────────────────────────────
# Bus#, Name (12 chars), kV, Area, Zone, Pgen, PMin, PMax
# Boston-area zones (per config.BOSTON_ZONES): 1000, 1002, 1004, 1010
# SEABROOK and NYPA excluded by config.EXCLUDED_PLANTS

GENERATORS = [
    [800001, "SAMPLE_BESS ", 115,   1, 1000,    0,   0,  100],  # project POI
    [800002, "ALPHA_GAS_1 ", 13.8,  1, 1000,  150,   0,  200],
    [800003, "ALPHA_GAS_2 ", 13.8,  1, 1000,  120,   0,  180],
    [800004, "BETA_GAS_A  ", 13.8,  1, 1002,   90,   0,  135],
    [800005, "BETA_GAS_B  ", 13.8,  1, 1002,   85,   0,  130],
    [800006, "GAMMA_JET   ", 13.8,  1, 1004,   45,   0,   70],
    [800007, "DELTA_GAS   ", 13.8,  1, 1004,   60,   0,   95],
    [800008, "EPSILON_WIND", 0.69,  1, 1010,   25,   0,   25],
    [800009, "ZETA_SOLAR  ", 0.69,  1, 1010,   18,   0,   18],
    [800010, "ETA_HYDRO   ", 13.8,  1, 1010,   40,   0,   60],
    [800011, "THETA_GAS   ", 13.8,  1, 4000,   80,   0,  120],
    [800012, "IOTA_GAS    ", 13.8,  1, 4000,   70,   0,  110],
    [800013, "KAPPA_WIND  ", 0.69,  1, 4000,   35,   0,   35],
    [800014, "LAMBDA_GAS  ", 13.8,  1, 1400,   55,   0,   85],
    [800015, "MU_SOLAR    ", 0.69,  1, 1400,   20,   0,   20],
    [800016, "NU_GAS      ", 13.8,  1, 1400,   65,   0,   95],
    [800017, "XI_WIND     ", 0.69,  1, 1630,   28,   0,   28],
    [800018, "OMICRON_GAS ", 13.8,  1, 1633,   75,   0,  110],
    [800019, "PI_GAS      ", 13.8,  1, 1633,   50,   0,   75],
    [800020, "RHO_HYDRO   ", 13.8,  1, 1634,   30,   0,   45],
    [999901, "SEABROOK_G1 ",  22,   1, 4000, 1200,   0, 1200],  # nuclear — excluded
    [999902, "NYPA_IMPORT ", 345,   1, 4000,  500,   0,  600],  # import  — excluded
]

# ─── Styles ──────────────────────────────────────────────────────────────────
BASE_FONT = Font(name="Courier New", size=10)
BOLD = Font(name="Courier New", size=10, bold=True)
RED_FONT = Font(name="Courier New", size=10, color="FF0000", bold=True)
FLOW_FILL = PatternFill("solid", fgColor="CCCCFF")
EDIT_FILL = PatternFill("solid", fgColor="FFFF00")
FORMULA_FILL = PatternFill("solid", fgColor="C0C0C0")
PNEW_FILL = PatternFill("solid", fgColor="BFBFBF")
THIN_SIDE = Side(style="thin", color="000000")
MEDIUM_SIDE = Side(style="medium", color="000000")


def apply_table_border(ws, min_row, max_row, min_col, max_col):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = Border(
                left=MEDIUM_SIDE if col == min_col else THIN_SIDE,
                right=MEDIUM_SIDE if col == max_col else THIN_SIDE,
                top=MEDIUM_SIDE if row == min_row else THIN_SIDE,
                bottom=MEDIUM_SIDE if row == max_row else THIN_SIDE,
            )


def write_sheet(wb, sheet_name, flows, dfax_table):
    """
    flows      : list of 5 flow rows - each is
                 [NN, None, element_str, None×5, contingency, limit, flowInit]
    dfax_table : list of 22 rows, each with 5 Dfax values
    """
    ws = wb.create_sheet(sheet_name)
    n_gens = len(GENERATORS)
    DATA_ROW = 14                       # first dispatch data row (Excel 1-indexed)
    HDR_ROW = DATA_ROW - 1              # dispatch header row
    SUM_ROW = HDR_ROW - 1               # "Dispatch" + SUM formulas row
    FLOW_HDR = 4                        # flow column-header row
    N_DFAX = 5

    for col in range(1, len(DISPATCH_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions["A"].width = 8.5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13

    for row in range(1, DATA_ROW + n_gens):
        ws.row_dimensions[row].height = 13.5
    ws.row_dimensions[3].height = 14
    ws.row_dimensions[12].height = 14
    ws.row_dimensions[13].height = 14

    ws.freeze_panes = "A14"

    # ── Row 1: comment header (pd.read_excel uses this as the column header) ──
    ws.cell(row=1, column=1).value = (
        "****** Plant Redispatch calculator ******* !!!! Please - edit only "
        "columns 'Status' & 'Pnew' !!!!!***"
    )
    ws.cell(row=1, column=1).font = BOLD

    # ── Row 3: "Flows" marker ─────────────────────────────────────────────────
    ws.cell(row=3, column=1).value = "Flows"
    ws.cell(row=3, column=1).font = BOLD

    # ── Row 4: flow column headers ────────────────────────────────────────────
    for col, val in enumerate(FLOW_HEADERS, 1):
        c = ws.cell(row=FLOW_HDR, column=col)
        c.value = val
        c.font = BOLD
        c.fill = FLOW_FILL

    # ── Rows 5-9: flow data ───────────────────────────────────────────────────
    for i, flow in enumerate(flows):
        r = FLOW_HDR + 1 + i          # Excel rows 5-9
        nn, _, element, *rest, contingency, limit, flow_init = flow

        ws.cell(row=r, column=1).value  = nn
        ws.cell(row=r, column=3).value  = element
        ws.cell(row=r, column=9).value  = contingency
        ws.cell(row=r, column=10).value = limit
        ws.cell(row=r, column=11).value = flow_init
        ws.cell(row=r, column=12).value = f"='{sheet_name}'!K{r}+'{sheet_name}'!M{r}"
        ws.cell(row=r, column=13).value = f"={get_column_letter(18 + i)}{SUM_ROW}"
        ws.cell(row=r, column=14).value = f"='{sheet_name}'!L{r}/'{sheet_name}'!J{r}"

        for col in range(1, 15):
            ws.cell(row=r, column=col).fill = FLOW_FILL
            ws.cell(row=r, column=col).font = BASE_FONT

        if flow_init / limit > 0.95:
            for col in [3, 9, 10, 11, 12, 13, 14]:
                ws.cell(row=r, column=col).font = RED_FONT

    apply_table_border(ws, FLOW_HDR, FLOW_HDR + len(flows), 1, 14)

    # ── Rows 10-11: empty ─────────────────────────────────────────────────────

    # ── Row 12 (SUM_ROW): "Dispatch" + column sums ───────────────────────────
    ws.cell(row=SUM_ROW, column=1).value = "Dispatch"
    ws.cell(row=SUM_ROW, column=1).font = BOLD

    # SUM of Pdelta (col J=10) and Impact_1-5 (cols R-V = 18-22)
    last = DATA_ROW + n_gens - 1
    ws.cell(row=SUM_ROW, column=10).value = f"=SUM(J{DATA_ROW}:J{last})"
    for imp_col in range(18, 18 + N_DFAX):
        ws.cell(row=SUM_ROW, column=imp_col).value = (
            f"=SUM({get_column_letter(imp_col)}{DATA_ROW}:"
            f"{get_column_letter(imp_col)}{last})"
        )

    # ── Row 13 (HDR_ROW): dispatch column headers + autofilter ───────────────
    for col, val in enumerate(DISPATCH_HEADERS, 1):
        c = ws.cell(row=HDR_ROW, column=col)
        c.value = val
        c.font = BOLD
        if col == 9:
            c.fill = EDIT_FILL
        elif col >= 10:
            c.fill = FORMULA_FILL
        else:
            c.fill = FLOW_FILL

    ws.auto_filter.ref = (
        f"A{HDR_ROW}:{get_column_letter(len(DISPATCH_HEADERS))}{last}"
    )

    apply_table_border(ws, HDR_ROW, last, 1, len(DISPATCH_HEADERS))

    # ── Rows 14+: generator dispatch data ────────────────────────────────────
    for idx, (gen, dfax) in enumerate(zip(GENERATORS, dfax_table)):
        r = DATA_ROW + idx
        bus, name, kv, area, zone, pgen, pmin, pmax = gen

        ws.cell(row=r, column=1).value  = bus
        ws.cell(row=r, column=2).value  = name
        ws.cell(row=r, column=3).value  = kv
        ws.cell(row=r, column=4).value  = area
        ws.cell(row=r, column=5).value  = zone
        ws.cell(row=r, column=6).value  = pgen
        ws.cell(row=r, column=7).value  = pmin
        ws.cell(row=r, column=8).value  = pmax
        ws.cell(row=r, column=9).value  = pgen          # Pnew = Pgen initially
        ws.cell(row=r, column=10).value = f"=I{r}-F{r}"
        ws.cell(row=r, column=11).value = f"=H{r}-I{r}"

        for col in range(1, len(DISPATCH_HEADERS) + 1):
            ws.cell(row=r, column=col).font = BASE_FONT

        # Dfax_1 through Dfax_5
        for k, d in enumerate(dfax[:N_DFAX], 12):
            ws.cell(row=r, column=k).value = d

        # --- separator
        ws.cell(row=r, column=17).value = None

        # Impact_1 through Impact_5 = Dfax_k * Pdelta
        for k, imp_col in enumerate(range(18, 18 + N_DFAX)):
            dfax_col = get_column_letter(12 + k)
            ws.cell(row=r, column=imp_col).value = f"={dfax_col}{r}*J{r}"

        ws.cell(row=r, column=9).fill = PNEW_FILL
        ws.cell(row=r, column=9).font = BOLD

    return ws


def main():
    # ── DFAX tables ───────────────────────────────────────────────────────────
    # 22 rows x 5 columns. Sorted by Dfax_1 descending (as the tool expects).
    # Boston-zone generators have higher positive DFAX on Boston import flows.

    sheet1_dfax = [
        # Dfax_1   Dfax_2   Dfax_3   Dfax_4   Dfax_5
        [ 0.523,   0.312,   0.287,   0.156,   0.098,   0.091],  # 800001 SAMPLE_BESS
        [ 0.445,   0.265,   0.241,   0.132,   0.082,   0.077],  # 800002 ALPHA_GAS_1
        [ 0.421,   0.251,   0.228,   0.124,   0.077,   0.073],  # 800003 ALPHA_GAS_2
        [ 0.384,   0.229,   0.208,   0.113,   0.070,   0.066],  # 800004 BETA_GAS_A
        [ 0.372,   0.222,   0.201,   0.109,   0.068,   0.064],  # 800005 BETA_GAS_B
        [ 0.298,   0.178,   0.161,   0.088,   0.054,   0.052],  # 800006 GAMMA_JET
        [ 0.276,   0.165,   0.149,   0.081,   0.050,   0.048],  # 800007 DELTA_GAS
        [ 0.198,   0.118,   0.107,   0.058,   0.036,   0.034],  # 800008 EPSILON_WIND
        [ 0.187,   0.112,   0.101,   0.055,   0.034,   0.032],  # 800009 ZETA_SOLAR
        [ 0.175,   0.105,   0.095,   0.051,   0.032,   0.030],  # 800010 ETA_HYDRO
        [ 0.082,   0.049,   0.044,   0.024,   0.015,   0.014],  # 800011 THETA_GAS
        [ 0.071,   0.042,   0.038,   0.021,   0.013,   0.012],  # 800012 IOTA_GAS
        [ 0.055,   0.033,   0.030,   0.016,   0.010,   0.009],  # 800013 KAPPA_WIND
        [ 0.041,   0.024,   0.022,   0.012,   0.007,   0.007],  # 800014 LAMBDA_GAS
        [ 0.033,   0.020,   0.018,   0.010,   0.006,   0.006],  # 800015 MU_SOLAR
        [ 0.028,   0.017,   0.015,   0.008,   0.005,   0.005],  # 800016 NU_GAS
        [ 0.019,   0.011,   0.010,   0.006,   0.004,   0.003],  # 800017 XI_WIND
        [ 0.014,   0.008,   0.008,   0.004,   0.003,   0.003],  # 800018 OMICRON_GAS
        [ 0.011,   0.007,   0.006,   0.003,   0.002,   0.002],  # 800019 PI_GAS
        [ 0.008,   0.005,   0.004,   0.002,   0.001,   0.001],  # 800020 RHO_HYDRO
        [-0.045,  -0.027,  -0.024,  -0.013,  -0.008,  -0.008],  # 999901 SEABROOK_G1
        [-0.098,  -0.058,  -0.053,  -0.029,  -0.018,  -0.017],  # 999902 NYPA_IMPORT
    ]

    sheet2_dfax = [
        [ 0.351,   0.198,   0.312,   0.087,   0.143,   0.151],
        [ 0.298,   0.168,   0.264,   0.074,   0.121,   0.128],
        [ 0.281,   0.159,   0.250,   0.070,   0.114,   0.121],
        [ 0.256,   0.145,   0.228,   0.063,   0.104,   0.110],
        [ 0.248,   0.140,   0.221,   0.061,   0.101,   0.106],
        [ 0.199,   0.112,   0.177,   0.049,   0.081,   0.085],
        [ 0.184,   0.104,   0.163,   0.046,   0.075,   0.079],
        [ 0.132,   0.074,   0.117,   0.033,   0.054,   0.057],
        [ 0.125,   0.070,   0.111,   0.031,   0.051,   0.054],
        [ 0.117,   0.066,   0.104,   0.029,   0.048,   0.050],
        [ 0.055,   0.031,   0.049,   0.014,   0.022,   0.024],
        [ 0.047,   0.027,   0.042,   0.012,   0.019,   0.020],
        [ 0.037,   0.021,   0.033,   0.009,   0.015,   0.016],
        [ 0.027,   0.015,   0.024,   0.007,   0.011,   0.012],
        [ 0.022,   0.012,   0.020,   0.005,   0.009,   0.010],
        [ 0.019,   0.010,   0.017,   0.005,   0.008,   0.008],
        [ 0.013,   0.007,   0.011,   0.003,   0.005,   0.006],
        [ 0.009,   0.005,   0.008,   0.002,   0.004,   0.004],
        [ 0.007,   0.004,   0.006,   0.002,   0.003,   0.003],
        [ 0.005,   0.003,   0.005,   0.001,   0.002,   0.002],
        [-0.030,  -0.017,  -0.027,  -0.007,  -0.012,  -0.013],
        [-0.065,  -0.037,  -0.058,  -0.016,  -0.027,  -0.028],
    ]

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = yxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: ALPHA_SUB BETA_SUB LINE_TRIP ────────────────────────────────
    # Flow 1 initial loading = 490/500 = 0.980.
    # Adding 50 MW (Dfax=0.523) → ΔFlow = 26.15 → FlowRes=516.15, Loading=1.032.
    # Redispatch reduces ALPHA_GAS_1 by ~50 MW → loading returns to ~0.987.
    sheet1_flows = [
        [1, None, "ALPHA_SUBS     115  BETA_SUBS      115  1 ",
         None, None, None, None, None, "LINE_ALP-BET_TRIP", 500, 490],
        [2, None, "GAMMA_SUBS     115  DELTA_SUBS     115  1 ",
         None, None, None, None, None, "LINE_ALP-BET_TRIP", 350, 280],
        [3, None, "EPSILON_SUBS   115  ZETA_SUBS      115  1 ",
         None, None, None, None, None, "LINE_ALP-BET_TRIP", 400, 360],
        [4, None, "ETA_SUBS       115  THETA_SUBS     115  1 ",
         None, None, None, None, None, "LINE_ALP-BET_TRIP", 300, 240],
        [5, None, "IOTA_SUBS      115  KAPPA_SUBS     115  1 ",
         None, None, None, None, None, "LINE_ALP-BET_TRIP", 600, 500],
    ]
    write_sheet(wb, "PRDcalc (2)", sheet1_flows, sheet1_dfax)

    # ── Sheet 2: GAMMA_SUB DELTA_SUB LINE_TRIP ───────────────────────────────
    # Flow 1 initial loading = 385/400 = 0.9625.
    # Adding 50 MW (Dfax=0.351) → ΔFlow = 17.55 → FlowRes=402.55, Loading=1.006.
    # Already within 1.02 limit — no intermediate file generated.
    sheet2_flows = [
        [1, None, "GAMMA_SUBS     115  DELTA_SUBS     115  1 ",
         None, None, None, None, None, "LINE_GAM-DEL_TRIP", 400, 385],
        [2, None, "ETA_SUBS       115  THETA_SUBS     115  1 ",
         None, None, None, None, None, "LINE_GAM-DEL_TRIP", 280, 220],
        [3, None, "ALPHA_SUBS     115  BETA_SUBS      115  1 ",
         None, None, None, None, None, "LINE_GAM-DEL_TRIP", 500, 420],
        [4, None, "LAMBDA_SUBS    115  MU_SUBS        115  1 ",
         None, None, None, None, None, "LINE_GAM-DEL_TRIP", 350, 270],
        [5, None, "NU_SUBS        115  XI_SUBS        115  1 ",
         None, None, None, None, None, "LINE_GAM-DEL_TRIP", 450, 380],
    ]
    write_sheet(wb, "PRDcalc", sheet2_flows, sheet2_dfax)

    out = os.path.join(os.path.dirname(__file__), "Sample_N-1_Study.xlsx")
    wb.save(out)
    print(f"Created: {out}")
    print()
    print("Suggested tool inputs:")
    print("  Bus Number  : 800001")
    print("  Project Size: 50")
    print("  Mode        : Discharging")


if __name__ == "__main__":
    main()
